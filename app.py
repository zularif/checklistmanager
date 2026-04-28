import sqlite3
import calendar
import os
from datetime import datetime, date
from flask import Flask, render_template, redirect, url_for, request, flash, send_file
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import pytz

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "checklist-secret-key-change-me")

# Always store the database next to this file, regardless of working directory
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE = os.path.join(BASE_DIR, "checklist.db")
MALAYSIA_TZ = pytz.timezone("Asia/Kuala_Lumpur")


def get_malaysia_now():
    return datetime.now(MALAYSIA_TZ)


def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


CHECKLIST_TASKS = [
    # ── Phase 1 ─────────────────────────────────────────────────
    # Daily
    ("Consumer Room HT 33 kV P1", "Daily"),
    ("Transformer Room 33 kV P1", "Daily"),
    ("MSB LV Room 420V P1", "Daily"),
    ("L-11 Testing Bay P1", "Daily"),
    ("Chiller Room P1", "Daily"),
    ("AHU & ACPU Room P1", "Daily"),
    ("Office FCU P1", "Daily"),
    ("Solar Panel P1", "Daily"),
    ("Production Line P1", "Daily"),
    ("Production (Hygrometer) P1", "Daily"),
    # Weekly
    ("Genset 415 VAC P1", "Weekly"),
    ("Solar String & Inverter P1", "Weekly"),
    ("ABB Type 2 EV Charging P1", "Weekly"),
    # Monthly
    ("Battery & Charger P1", "Monthly"),
    ("Capacitor Bank (MSB Room) P1", "Monthly"),
    ("ESD Earth P1", "Monthly"),
    ("Water Meter Reading P1", "Monthly"),
    ("Cargo Lift Inspection P1", "Monthly"),
    ("Carbon Filter Checklist P1", "Monthly"),
    ("Carbon Pellet P1", "Monthly"),
    ("Grease Filter P1", "Monthly"),
    ("AHU Panel P1", "Monthly"),
    # ── Phase 2 ─────────────────────────────────────────────────
    # Daily
    ("Consumer Room HT 33 kV P2", "Daily"),
    ("Transformer Room 33 kV P2", "Daily"),
    ("MSB LV Room 420V P2", "Daily"),
    ("Solar Panel (2) P2", "Daily"),
    ("Chiller Room P2", "Daily"),
    ("Production Line 1 P2", "Daily"),
    ("Production Line 2 P2", "Daily"),
    ("Production (Hygrometer) P2", "Daily"),
    # Weekly
    ("Genset 415 VAC P2", "Weekly"),
    ("Solar String & Inverter P2", "Weekly"),
    ("ABB Type 2 EV Charging P2", "Weekly"),
    # Monthly
    ("Battery & Charger P2", "Monthly"),
    ("Air Particles Measurement Record P2", "Monthly"),
    ("Capacitor Bank (MSB Room) P2", "Monthly"),
    ("ESD Earth P2", "Monthly"),
    ("Water Meter Reading P2", "Monthly"),
    ("Cargo Lift Inspection P2", "Monthly"),
    ("Carbon Filter Checklist P2", "Monthly"),
    ("Carbon Pellet P2", "Monthly"),
    ("Grease Filter P2", "Monthly"),
    ("AHU Panel P2", "Monthly"),
]


ALL_FREQUENCIES = [
    "Daily",
    "Weekly",
    "Monthly",
    "Quarterly",
    "Semi-Annual",
    "Annual",
    "Bi-Annual",
]


def init_db():
    conn = get_db()
    cursor = conn.cursor()

    cursor.executescript("""
        CREATE TABLE IF NOT EXISTS tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            location TEXT NOT NULL,
            frequency TEXT NOT NULL,
            type TEXT NOT NULL DEFAULT 'Internal',
            vendor TEXT
        );

        CREATE TABLE IF NOT EXISTS task_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id INTEGER NOT NULL,
            completed_date TEXT NOT NULL,
            technician TEXT NOT NULL DEFAULT '',
            FOREIGN KEY (task_id) REFERENCES tasks(id)
        );

        CREATE TABLE IF NOT EXISTS common_tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            description TEXT,
            priority TEXT NOT NULL DEFAULT 'Medium' CHECK(priority IN ('High', 'Medium', 'Low')),
            status TEXT NOT NULL DEFAULT 'Open' CHECK(status IN ('Open', 'In Progress', 'Done')),
            created_date TEXT NOT NULL,
            completed_date TEXT
        );
    """)
    conn.commit()

    # Migration: add technician to task_logs if missing
    log_cols = [
        row[1] for row in cursor.execute("PRAGMA table_info(task_logs)").fetchall()
    ]
    if "technician" not in log_cols:
        cursor.execute(
            "ALTER TABLE task_logs ADD COLUMN technician TEXT NOT NULL DEFAULT ''"
        )
        conn.commit()

    # Migration: add type/vendor to tasks + remove old frequency CHECK constraint
    task_cols = [
        row[1] for row in cursor.execute("PRAGMA table_info(tasks)").fetchall()
    ]
    if "type" not in task_cols:
        # Recreate tasks table without old CHECK, adding new columns
        cursor.executescript("""
            CREATE TABLE tasks_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                location TEXT NOT NULL,
                frequency TEXT NOT NULL,
                type TEXT NOT NULL DEFAULT 'Internal',
                vendor TEXT
            );
            INSERT INTO tasks_new (id, location, frequency)
                SELECT id, location, frequency FROM tasks;
            DROP TABLE tasks;
            ALTER TABLE tasks_new RENAME TO tasks;
        """)
        conn.commit()

    cursor.execute("SELECT COUNT(*) FROM tasks")
    if cursor.fetchone()[0] == 0:
        cursor.executemany(
            "INSERT INTO tasks (location, frequency) VALUES (?, ?)", CHECKLIST_TASKS
        )
        conn.commit()

    conn.close()


# ─────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────


def get_month_completion_count(task_id, year, month):
    conn = get_db()
    cursor = conn.cursor()
    month_prefix = f"{year}-{month:02d}"
    cursor.execute(
        "SELECT COUNT(*) FROM task_logs WHERE task_id = ? AND completed_date LIKE ?",
        (task_id, f"{month_prefix}%"),
    )
    count = cursor.fetchone()[0]
    conn.close()
    return count


def get_period_count(task_id, frequency, year, month):
    """Returns completion count for the relevant period based on frequency."""
    conn = get_db()
    cursor = conn.cursor()
    if frequency in ("Daily", "Weekly", "Monthly"):
        prefix = f"{year}-{month:02d}"
        cursor.execute(
            "SELECT COUNT(*) FROM task_logs WHERE task_id=? AND completed_date LIKE ?",
            (task_id, f"{prefix}%"),
        )
    elif frequency == "Quarterly":
        q = (month - 1) // 3
        sm = q * 3 + 1
        em = sm + 3
        start = f"{year}-{sm:02d}-01"
        end = f"{year}-{em:02d}-01" if em <= 12 else f"{year + 1}-01-01"
        cursor.execute(
            "SELECT COUNT(*) FROM task_logs WHERE task_id=? AND completed_date>=? AND completed_date<?",
            (task_id, start, end),
        )
    elif frequency == "Semi-Annual":
        sm = 1 if month <= 6 else 7
        em = 7 if month <= 6 else 13
        start = f"{year}-{sm:02d}-01"
        end = f"{year}-{em:02d}-01" if em <= 12 else f"{year + 1}-01-01"
        cursor.execute(
            "SELECT COUNT(*) FROM task_logs WHERE task_id=? AND completed_date>=? AND completed_date<?",
            (task_id, start, end),
        )
    elif frequency == "Annual":
        cursor.execute(
            "SELECT COUNT(*) FROM task_logs WHERE task_id=? AND completed_date>=? AND completed_date<?",
            (task_id, f"{year}-01-01", f"{year + 1}-01-01"),
        )
    elif frequency == "Bi-Annual":
        cursor.execute(
            "SELECT COUNT(*) FROM task_logs WHERE task_id=? AND completed_date>=? AND completed_date<?",
            (task_id, f"{year - 1}-01-01", f"{year + 1}-01-01"),
        )
    else:
        prefix = f"{year}-{month:02d}"
        cursor.execute(
            "SELECT COUNT(*) FROM task_logs WHERE task_id=? AND completed_date LIKE ?",
            (task_id, f"{prefix}%"),
        )
    count = cursor.fetchone()[0]
    conn.close()
    return count


def get_period_label(frequency, year, month):
    if frequency == "Quarterly":
        q = (month - 1) // 3 + 1
        return f"Q{q} {year}"
    elif frequency == "Semi-Annual":
        h = 1 if month <= 6 else 2
        return f"H{h} {year}"
    elif frequency == "Annual":
        return str(year)
    elif frequency == "Bi-Annual":
        return f"{year - 1}–{year}"
    return ""


def is_task_complete(frequency, count, year, month):
    if frequency == "Daily":
        return count >= calendar.monthrange(year, month)[1]
    elif frequency == "Weekly":
        return count >= 4
    elif frequency in ("Monthly", "Quarterly", "Semi-Annual", "Annual", "Bi-Annual"):
        return count >= 1
    return False


# ─────────────────────────────────────────────────────────────────
# Checklist routes
# ─────────────────────────────────────────────────────────────────


@app.route("/")
def task_list():
    now = get_malaysia_now()
    year, month = now.year, now.month

    conn = get_db()
    tasks = conn.execute("SELECT * FROM tasks ORDER BY frequency, location").fetchall()
    conn.close()

    phase_groups = {f: {"P1": [], "P2": []} for f in ALL_FREQUENCIES}

    for task in tasks:
        freq = task["frequency"]
        if freq not in phase_groups:
            phase_groups[freq] = {"P1": [], "P2": []}
        count = get_period_count(task["id"], freq, year, month)
        complete = is_task_complete(freq, count, year, month)
        entry = {
            "id": task["id"],
            "location": task["location"],
            "frequency": freq,
            "type": task["type"] if task["type"] else "Internal",
            "vendor": task["vendor"] or "",
            "count": count,
            "complete": complete,
            "period_label": get_period_label(freq, year, month),
        }
        phase = "P1" if "P1" in task["location"] else "P2"
        phase_groups[freq][phase].append(entry)

    default_dt = now.strftime("%Y-%m-%dT%H:%M")
    return render_template(
        "task_list.html",
        phase_groups=phase_groups,
        all_frequencies=ALL_FREQUENCIES,
        month_name=now.strftime("%B %Y"),
        default_dt=default_dt,
    )


@app.route("/complete/<int:task_id>", methods=["POST"])
def complete_task(task_id):
    technician = request.form.get("technician", "").strip()
    if not technician:
        flash("Please enter the technician name before marking complete.", "error")
        return redirect(url_for("task_list"))

    raw_date = request.form.get("completed_date", "").strip()
    if raw_date:
        try:
            # datetime-local format: YYYY-MM-DDTHH:MM
            dt = datetime.strptime(raw_date, "%Y-%m-%dT%H:%M")
            completed_date = dt.strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            flash("Invalid date format. Using current time.", "error")
            completed_date = get_malaysia_now().strftime("%Y-%m-%d %H:%M:%S")
    else:
        completed_date = get_malaysia_now().strftime("%Y-%m-%d %H:%M:%S")

    conn = get_db()
    task = conn.execute("SELECT id FROM tasks WHERE id = ?", (task_id,)).fetchone()
    if task:
        conn.execute(
            "INSERT INTO task_logs (task_id, completed_date, technician) VALUES (?, ?, ?)",
            (task_id, completed_date, technician),
        )
        conn.commit()
        flash(
            f"Completed by {technician} — recorded at {completed_date[:16]}.", "success"
        )
    else:
        flash("Task not found.", "error")
    conn.close()
    return redirect(url_for("task_list"))


@app.route("/history")
def history():
    month_filter = request.args.get("month", "")
    tech_filter = request.args.get("tech", "").strip()

    now = get_malaysia_now()
    default_month = now.strftime("%Y-%m")
    if not month_filter:
        month_filter = default_month

    conn = get_db()

    query = """
        SELECT tl.id, tl.completed_date, tl.technician,
               t.location, t.frequency
        FROM task_logs tl
        JOIN tasks t ON tl.task_id = t.id
        WHERE tl.completed_date LIKE ?
    """
    params = [f"{month_filter}%"]

    if tech_filter:
        query += " AND LOWER(tl.technician) LIKE ?"
        params.append(f"%{tech_filter.lower()}%")

    query += " ORDER BY tl.completed_date DESC"
    logs = conn.execute(query, params).fetchall()

    # All distinct technicians for this month (for filter dropdown)
    techs = conn.execute(
        "SELECT DISTINCT technician FROM task_logs WHERE completed_date LIKE ? ORDER BY technician",
        (f"{month_filter}%",),
    ).fetchall()

    # Summary: per technician count this month
    summary = conn.execute(
        """SELECT tl.technician, COUNT(*) as cnt
           FROM task_logs tl
           WHERE tl.completed_date LIKE ?
           GROUP BY tl.technician
           ORDER BY cnt DESC""",
        (f"{month_filter}%",),
    ).fetchall()

    conn.close()

    month_label = datetime.strptime(month_filter, "%Y-%m").strftime("%B %Y")
    return render_template(
        "history.html",
        logs=logs,
        summary=summary,
        techs=techs,
        month_filter=month_filter,
        tech_filter=tech_filter,
        month_label=month_label,
    )


@app.route("/dashboard")
def dashboard():
    now = get_malaysia_now()
    current_year = now.year
    current_month = now.month

    try:
        selected_year = int(request.args.get("year", current_year))
    except ValueError:
        selected_year = current_year

    MONTHS = [
        "JAN",
        "FEB",
        "MAR",
        "APR",
        "MAY",
        "JUN",
        "JUL",
        "AUG",
        "SEP",
        "OCT",
        "NOV",
        "DEC",
    ]

    conn = get_db()
    tasks = conn.execute("SELECT * FROM tasks ORDER BY frequency, location").fetchall()

    from collections import defaultdict

    def _fetch_log_map(yr):
        rows = conn.execute(
            "SELECT task_id, completed_date, technician FROM task_logs "
            "WHERE completed_date LIKE ? ORDER BY completed_date",
            (f"{yr}-%",),
        ).fetchall()
        lm = defaultdict(lambda: defaultdict(list))
        for r in rows:
            mo = int(r["completed_date"][5:7])
            lm[r["task_id"]][mo].append(r["technician"] or "—")
        return lm

    log_map = _fetch_log_map(selected_year)
    prev_log_map = _fetch_log_map(selected_year - 1)
    conn.close()

    def _build_cell(task_id, freq, m):
        """Return one dashboard cell dict for (task_id, frequency, month m)."""
        if freq == "Daily":
            count = len(log_map[task_id][m])
            required = calendar.monthrange(selected_year, m)[1]
            techs = log_map[task_id][m]
            is_future = (
                selected_year == current_year and m > current_month
            ) or selected_year > current_year
        elif freq == "Weekly":
            count = len(log_map[task_id][m])
            required = 4
            techs = log_map[task_id][m]
            is_future = (
                selected_year == current_year and m > current_month
            ) or selected_year > current_year
        elif freq == "Monthly":
            count = len(log_map[task_id][m])
            required = 1
            techs = log_map[task_id][m]
            is_future = (
                selected_year == current_year and m > current_month
            ) or selected_year > current_year
        elif freq == "Quarterly":
            q_idx = (m - 1) // 3
            sm = q_idx * 3 + 1
            em = sm + 3
            count = sum(len(log_map[task_id][i]) for i in range(sm, em))
            techs = [t for i in range(sm, em) for t in log_map[task_id][i]]
            required = 1
            is_future = (
                selected_year == current_year and sm > current_month
            ) or selected_year > current_year
        elif freq == "Semi-Annual":
            sm = 1 if m <= 6 else 7
            em = 7 if m <= 6 else 13
            count = sum(len(log_map[task_id][i]) for i in range(sm, em))
            techs = [t for i in range(sm, em) for t in log_map[task_id][i]]
            required = 1
            is_future = (
                selected_year == current_year and sm > current_month
            ) or selected_year > current_year
        elif freq == "Annual":
            count = sum(len(log_map[task_id][i]) for i in range(1, 13))
            techs = [t for i in range(1, 13) for t in log_map[task_id][i]]
            required = 1
            is_future = selected_year > current_year
        elif freq == "Bi-Annual":
            count = sum(len(log_map[task_id][i]) for i in range(1, 13)) + sum(
                len(prev_log_map[task_id][i]) for i in range(1, 13)
            )
            techs = [t for i in range(1, 13) for t in log_map[task_id][i]]
            required = 1
            is_future = selected_year > current_year
        else:
            count, required, techs, is_future = 0, 1, [], False

        return {
            "count": count,
            "required": required,
            "complete": count >= required,
            "future": is_future,
            "techs": techs,
        }

    sections = []
    for freq in ALL_FREQUENCIES:
        p1_rows, p2_rows = [], []
        for task in tasks:
            if task["frequency"] != freq:
                continue
            cells = [_build_cell(task["id"], freq, m) for m in range(1, 13)]
            row = {
                "location": task["location"],
                "frequency": freq,
                "type": task["type"] or "Internal",
                "vendor": task["vendor"] or "",
                "cells": cells,
            }
            if "P1" in task["location"]:
                p1_rows.append(row)
            else:
                p2_rows.append(row)
        if p1_rows or p2_rows:
            sections.append({"frequency": freq, "p1": p1_rows, "p2": p2_rows})

    year_range = list(range(current_year - 2, current_year + 2))
    return render_template(
        "dashboard.html",
        sections=sections,
        months=MONTHS,
        selected_year=selected_year,
        current_year=current_year,
        current_month=current_month,
        year_range=year_range,
    )


def _s(ws, row, col, value, alignment, font, border):
    """Helper: set cell value + style."""
    c = ws.cell(row=row, column=col)
    c.value = value
    c.alignment = alignment
    c.font = font
    c.border = border
    return c


@app.route("/export/<int:year>")
def export_excel(year):
    from collections import defaultdict

    conn = get_db()
    tasks = conn.execute("SELECT * FROM tasks ORDER BY frequency, location").fetchall()

    def _fetch_lm(yr):
        rows = conn.execute(
            "SELECT task_id, completed_date, technician FROM task_logs "
            "WHERE completed_date LIKE ? ORDER BY completed_date",
            (f"{yr}-%",),
        ).fetchall()
        lm = defaultdict(lambda: defaultdict(list))
        for r in rows:
            mo = int(r["completed_date"][5:7])
            lm[r["task_id"]][mo].append(r["technician"] or "—")
        return lm

    log_map = _fetch_lm(year)
    prev_log_map = _fetch_lm(year - 1)
    conn.close()

    current_now = get_malaysia_now()
    current_year = current_now.year
    current_month = current_now.month

    MONTHS = [
        "JAN",
        "FEB",
        "MAR",
        "APR",
        "MAY",
        "JUN",
        "JUL",
        "AUG",
        "SEP",
        "OCT",
        "NOV",
        "DEC",
    ]

    CYCLE_CODES = {
        "Daily": "D",
        "Weekly": "W",
        "Monthly": "M",
        "Quarterly": "Q",
        "Semi-Annual": "6M",
        "Annual": "1Y",
        "Bi-Annual": "2Y",
    }

    # ── Styles ──────────────────────────────────────────────────────
    fill_green = PatternFill("solid", fgColor="C6EFCE")
    fill_red = PatternFill("solid", fgColor="FFC7CE")
    fill_grey = PatternFill("solid", fgColor="EFEFEF")
    fill_header = PatternFill("solid", fgColor="1A3E5C")
    fill_phase1 = PatternFill("solid", fgColor="2A6496")
    fill_phase2 = PatternFill("solid", fgColor="276F6E")
    fill_summary = PatternFill("solid", fgColor="EBF4FF")
    fill_na = PatternFill("solid", fgColor="F8F8F8")

    SECTION_FILLS = {
        "Daily": PatternFill("solid", fgColor="D6E8F5"),
        "Weekly": PatternFill("solid", fgColor="D4EDDA"),
        "Monthly": PatternFill("solid", fgColor="FFF3CD"),
        "Quarterly": PatternFill("solid", fgColor="EDE0FF"),
        "Semi-Annual": PatternFill("solid", fgColor="FFE0E0"),
        "Annual": PatternFill("solid", fgColor="FFF0DC"),
        "Bi-Annual": PatternFill("solid", fgColor="DCE4FF"),
    }

    font_white_bold = Font(color="FFFFFF", bold=True, size=10)
    font_header = Font(color="FFFFFF", bold=True, size=10)
    font_normal = Font(size=9)
    font_section = Font(bold=True, size=10, color="1A3E5C")
    font_green = Font(color="276221", size=9)
    font_red = Font(color="9C0006", size=9)
    font_grey = Font(color="AAAAAA", size=9)
    font_summary = Font(bold=True, size=9, color="1A3E5C")

    thin_side = Side(style="thin", color="BFBFBF")
    thin_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )
    center_a = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_a = Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Inspection {year}"

    TOTAL_COLS = 17  # A=No, B=Location, C=Cycle, D-O=months, P=Done, Q=%

    # ── Main title ──────────────────────────────────────────────────
    ws.merge_cells("A1:Q1")
    tc = ws["A1"]
    tc.value = f"Factory Inspection Checklist — {year}"
    tc.fill = fill_header
    tc.font = Font(color="FFFFFF", bold=True, size=13)
    tc.alignment = center_a
    ws.row_dimensions[1].height = 28

    # ── Column headers ───────────────────────────────────────────────
    ws.append(["No.", "Location", "Cycle"] + MONTHS + ["Done", "% Rate"])
    hrow = 2
    for col in range(1, TOTAL_COLS + 1):
        c = ws.cell(row=hrow, column=col)
        c.fill = fill_header
        c.font = font_header
        c.alignment = center_a
        c.border = thin_border
    ws.row_dimensions[hrow].height = 20

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 7
    for ltr in ["D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]:
        ws.column_dimensions[ltr].width = 11
    ws.column_dimensions["P"].width = 7
    ws.column_dimensions["Q"].width = 9
    ws.freeze_panes = "D3"

    row_num = 3
    task_no = 1

    def _cell_data(task_id, freq, m):
        """Compute (count, required, is_future, techs) for one monthly cell."""
        if freq in ("Daily", "Weekly", "Monthly"):
            count = len(log_map[task_id][m])
            required = (
                calendar.monthrange(year, m)[1]
                if freq == "Daily"
                else 4
                if freq == "Weekly"
                else 1
            )
            techs = log_map[task_id][m]
            is_future = (
                year == current_year and m > current_month
            ) or year > current_year
        elif freq == "Quarterly":
            q_idx = (m - 1) // 3
            sm, em = q_idx * 3 + 1, q_idx * 3 + 4
            count = sum(len(log_map[task_id][i]) for i in range(sm, em))
            techs = [t for i in range(sm, em) for t in log_map[task_id][i]]
            required = 1
            is_future = (
                year == current_year and sm > current_month
            ) or year > current_year
        elif freq == "Semi-Annual":
            sm, em = (1, 7) if m <= 6 else (7, 13)
            count = sum(len(log_map[task_id][i]) for i in range(sm, em))
            techs = [t for i in range(sm, em) for t in log_map[task_id][i]]
            required = 1
            is_future = (
                year == current_year and sm > current_month
            ) or year > current_year
        elif freq == "Annual":
            count = sum(len(log_map[task_id][i]) for i in range(1, 13))
            techs = [t for i in range(1, 13) for t in log_map[task_id][i]]
            required = 1
            is_future = year > current_year
        elif freq == "Bi-Annual":
            count = sum(len(log_map[task_id][i]) for i in range(1, 13)) + sum(
                len(prev_log_map[task_id][i]) for i in range(1, 13)
            )
            techs = [t for i in range(1, 13) for t in log_map[task_id][i]]
            required = 1
            is_future = year > current_year
        else:
            count, required, techs, is_future = 0, 1, [], False
        return count, required, is_future, techs

    def _done_rate(task_id, freq):
        """Compute (done_periods, applicable_periods) for summary columns."""
        if freq in ("Daily", "Weekly", "Monthly"):
            return None, None  # use formula

        if freq == "Quarterly":
            periods = [((q * 3 + 1), (q * 3 + 4)) for q in range(4)]
        elif freq == "Semi-Annual":
            periods = [(1, 7), (7, 13)]
        elif freq in ("Annual", "Bi-Annual"):
            periods = [(1, 13)]
        else:
            return None, None

        done = 0
        applicable = 0
        for sm, em in periods:
            if freq == "Bi-Annual":
                cnt = sum(len(log_map[task_id][i]) for i in range(sm, em)) + sum(
                    len(prev_log_map[task_id][i]) for i in range(sm, em)
                )
            else:
                cnt = sum(len(log_map[task_id][i]) for i in range(sm, em))
            fut = (year == current_year and sm > current_month) or year > current_year
            if not fut:
                applicable += 1
                if cnt >= 1:
                    done += 1
        return done, applicable

    for freq in ALL_FREQUENCIES:
        freq_tasks = [t for t in tasks if t["frequency"] == freq]
        if not freq_tasks:
            continue

        # Section header
        ws.merge_cells(
            start_row=row_num, start_column=1, end_row=row_num, end_column=TOTAL_COLS
        )
        sc = ws.cell(row=row_num, column=1)
        sc.value = f"▶  {freq.upper()} TASKS"
        sc.fill = SECTION_FILLS.get(freq, PatternFill("solid", fgColor="D9E1F2"))
        sc.font = font_section
        sc.alignment = left_a
        sc.border = thin_border
        ws.row_dimensions[row_num].height = 18
        row_num += 1

        for phase_key, phase_fill, phase_label in [
            ("P1", fill_phase1, "Phase 1"),
            ("P2", fill_phase2, "Phase 2"),
        ]:
            phase_tasks = [t for t in freq_tasks if phase_key in t["location"]]
            if not phase_tasks:
                continue

            # Phase sub-header
            ws.merge_cells(
                start_row=row_num,
                start_column=1,
                end_row=row_num,
                end_column=TOTAL_COLS,
            )
            pc = ws.cell(row=row_num, column=1)
            pc.value = f"  {phase_label}"
            pc.fill = phase_fill
            pc.font = font_white_bold
            pc.alignment = left_a
            pc.border = thin_border
            ws.row_dimensions[row_num].height = 16
            row_num += 1

            for task in phase_tasks:
                r = row_num
                # Col A – No.
                _s(ws, r, 1, task_no, center_a, font_normal, thin_border)
                # Col B – Location (+ vendor if external)
                loc_val = task["location"]
                if task["vendor"]:
                    loc_val += f"\n[{task['vendor']}]"
                _s(ws, r, 2, loc_val, left_a, font_normal, thin_border)
                # Col C – Cycle code
                _s(
                    ws,
                    r,
                    3,
                    CYCLE_CODES.get(freq, freq[0]),
                    center_a,
                    font_normal,
                    thin_border,
                )

                # Cols D-O – month cells
                for m in range(1, 13):
                    col = m + 3
                    count, required, is_future, techs = _cell_data(task["id"], freq, m)
                    cell = ws.cell(row=r, column=col)
                    cell.alignment = center_a
                    cell.border = thin_border
                    if is_future:
                        cell.value = "—"
                        cell.fill = fill_grey
                        cell.font = font_grey
                    elif count >= required:
                        tech_str = ", ".join(sorted(set(techs))) if techs else ""
                        cell.value = f"✔\n{count}x" + (
                            f"\n{tech_str}" if tech_str else ""
                        )
                        cell.fill = fill_green
                        cell.font = font_green
                    else:
                        cell.value = f"✘\n{count}/{required}"
                        cell.fill = fill_red
                        cell.font = font_red

                # Col P – Done
                done_c = ws.cell(row=r, column=16)
                done_c.alignment = center_a
                done_c.border = thin_border
                done_c.font = font_summary
                done_c.fill = fill_summary

                # Col Q – % Rate
                rate_c = ws.cell(row=r, column=17)
                rate_c.alignment = center_a
                rate_c.border = thin_border
                rate_c.font = font_summary
                rate_c.fill = fill_summary
                rate_c.number_format = "0%"

                if freq in ("Daily", "Weekly", "Monthly"):
                    done_c.value = f'=COUNTIF(D{r}:O{r},"✔*")'
                    rate_c.value = (
                        f'=IF((12-COUNTIF(D{r}:O{r},"—"))>0,'
                        f'P{r}/(12-COUNTIF(D{r}:O{r},"—")),0)'
                    )
                else:
                    done_val, appl_val = _done_rate(task["id"], freq)
                    done_c.value = done_val if done_val is not None else "—"
                    if appl_val:
                        rate_c.value = done_val / appl_val
                    else:
                        rate_c.value = 0

                ws.row_dimensions[r].height = 38
                row_num += 1
                task_no += 1

    # ── Timestamp ────────────────────────────────────────────────────
    row_num += 1
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
    ts = ws.cell(row=row_num, column=1)
    ts.value = (
        f"Generated: {get_malaysia_now().strftime('%d %b %Y  %H:%M')} (Malaysia time)"
    )
    ts.font = Font(italic=True, size=8, color="7F7F7F")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"Factory_Inspection_{year}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/export/template")
def export_template():
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule

    now = get_malaysia_now()
    year = now.year

    conn = get_db()
    tasks = conn.execute("SELECT * FROM tasks ORDER BY frequency, location").fetchall()
    conn.close()

    MONTHS = [
        "JAN",
        "FEB",
        "MAR",
        "APR",
        "MAY",
        "JUN",
        "JUL",
        "AUG",
        "SEP",
        "OCT",
        "NOV",
        "DEC",
    ]

    wb = openpyxl.Workbook()

    # ── Shared styles ─────────────────────────────────────────────────
    thin_s = Side(style="thin", color="BFBFBF")
    med_s = Side(style="medium", color="1A3E5C")
    thin_b = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
    med_b = Border(left=med_s, right=med_s, top=med_s, bottom=med_s)
    fill_hdr = PatternFill("solid", fgColor="1A3E5C")
    fill_entry = PatternFill("solid", fgColor="FFFDE7")  # light amber = editable
    fill_blue = PatternFill("solid", fgColor="EBF4FF")
    fill_sec = PatternFill("solid", fgColor="D9E1F2")
    fill_p1 = PatternFill("solid", fgColor="2A6496")
    fill_p2 = PatternFill("solid", fgColor="276F6E")
    fill_green = PatternFill("solid", fgColor="C6EFCE")
    fill_red = PatternFill("solid", fgColor="FFC7CE")
    fill_grey = PatternFill("solid", fgColor="EFEFEF")
    fnt_hdr = Font(color="FFFFFF", bold=True, size=10)
    fnt_wb = Font(color="FFFFFF", bold=True, size=10)
    fnt_sec = Font(bold=True, size=10, color="1A3E5C")
    fnt_norm = Font(size=9)
    fnt_green = Font(color="276221", size=9, bold=True)
    fnt_red = Font(color="9C0006", size=9, bold=True)
    fnt_grey = Font(color="7F7F7F", size=9)
    fnt_blueb = Font(bold=True, size=9, color="1A3E5C")
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ════════════════════════════════════════════════════════════════
    # SHEET 1 — CHECKLIST  (entry + live dashboard combined)
    # ════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Checklist"

    # Row 1 — Main title
    ws.merge_cells("A1:R1")
    ws["A1"].value = "Factory Inspection Checklist"
    ws["A1"].fill = fill_hdr
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = ctr
    ws.row_dimensions[1].height = 30

    # Row 2 — Year selector
    ws["A2"].value = "YEAR"
    ws["A2"].font = Font(bold=True, size=11, color="1A3E5C")
    ws["A2"].alignment = Alignment(horizontal="right", vertical="center")
    ws["B2"].value = year
    ws["B2"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["B2"].fill = PatternFill("solid", fgColor="2B6CB0")
    ws["B2"].alignment = ctr
    ws["B2"].border = med_b
    ws["C2"].value = "  ← Change the year here (e.g. 2025)"
    ws["C2"].font = Font(italic=True, size=9, color="888888")
    ws["C2"].alignment = lft
    ws.row_dimensions[2].height = 24

    # Row 3 — Legend / instructions
    ws.merge_cells("A3:R3")
    ws["A3"].value = (
        "  HOW TO USE:  Find the task row and the month column.  "
        "Type the count in the cell (Daily = days done, Weekly = weeks done, Monthly = 1).  "
        "Cells turn GREEN when target is met, RED when not.  GREY = future month."
    )
    ws["A3"].font = Font(italic=True, size=9, color="444444")
    ws["A3"].fill = PatternFill("solid", fgColor="FFF9C4")
    ws["A3"].alignment = lft
    ws.row_dimensions[3].height = 18

    # Row 4 — Column headers
    HDR = 4
    ws.append(["No.", "Location", "F", "Target"] + MONTHS + ["Done", "%"])
    for c in range(1, 19):
        cell = ws.cell(row=HDR, column=c)
        cell.fill = fill_hdr
        cell.font = fnt_hdr
        cell.alignment = ctr
        cell.border = thin_b
    ws.row_dimensions[HDR].height = 20

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 13
    for L in ["E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]:
        ws.column_dimensions[L].width = 8
    ws.column_dimensions["Q"].width = 6
    ws.column_dimensions["R"].width = 7

    ws.freeze_panes = "E5"

    # ── Task rows ─────────────────────────────────────────────────────
    row_num = 5
    task_no = 1
    data_rows = []  # (row, freq_code)

    for freq in ["Daily", "Weekly", "Monthly"]:
        freq_tasks = [t for t in tasks if t["frequency"] == freq]
        if not freq_tasks:
            continue
        fc = freq[0]  # D / W / M
        if fc == "D":
            target_txt = "All days/mo"
        elif fc == "W":
            target_txt = "≥ 4 / month"
        else:
            target_txt = "≥ 1 / month"

        # Section header (spans all 18 cols)
        ws.merge_cells(
            start_row=row_num, start_column=1, end_row=row_num, end_column=18
        )
        sc = ws.cell(row=row_num, column=1)
        sc.value = f"▶  {freq.upper()} TASKS"
        sc.fill = fill_sec
        sc.font = fnt_sec
        sc.alignment = lft
        sc.border = thin_b
        ws.row_dimensions[row_num].height = 18
        row_num += 1

        for pk, pf, pl in [("P1", fill_p1, "Phase 1"), ("P2", fill_p2, "Phase 2")]:
            ptasks = [t for t in freq_tasks if pk in t["location"]]
            if not ptasks:
                continue

            # Phase sub-header
            ws.merge_cells(
                start_row=row_num, start_column=1, end_row=row_num, end_column=18
            )
            ph = ws.cell(row=row_num, column=1)
            ph.value = f"  {pl}"
            ph.fill = pf
            ph.font = fnt_wb
            ph.alignment = lft
            ph.border = thin_b
            ws.row_dimensions[row_num].height = 15
            row_num += 1

            for task in ptasks:
                r = row_num

                # No.
                c1 = ws.cell(row=r, column=1)
                c1.value = task_no
                c1.alignment = ctr
                c1.border = thin_b
                c1.font = fnt_norm

                # Location
                c2 = ws.cell(row=r, column=2)
                c2.value = task["location"]
                c2.alignment = lft
                c2.border = thin_b
                c2.font = fnt_norm

                # Freq code
                c3 = ws.cell(row=r, column=3)
                c3.value = fc
                c3.alignment = ctr
                c3.border = thin_b
                c3.font = fnt_norm

                # Target
                c4 = ws.cell(row=r, column=4)
                c4.value = target_txt
                c4.alignment = ctr
                c4.border = thin_b
                c4.font = Font(size=8, italic=True, color="555555")

                # Month entry cells  E(col5)=JAN … P(col16)=DEC
                for m in range(1, 13):
                    col = m + 4  # E=5=JAN … P=16=DEC
                    mc = ws.cell(row=r, column=col)
                    mc.value = None  # blank = user types here
                    mc.number_format = "0"
                    mc.alignment = ctr
                    mc.border = thin_b
                    mc.font = Font(size=10, bold=True)
                    mc.fill = fill_entry  # amber = editable

                # Done  (Q = col 17)
                # count months where entry meets requirement
                qc = ws.cell(row=r, column=17)
                qc.fill = fill_blue
                qc.alignment = ctr
                qc.border = thin_b
                qc.font = fnt_blueb
                if fc == "D":
                    # sum: 1 if count >= days in that month
                    done_parts = "+".join(
                        f"(ISNUMBER({get_column_letter(m + 4)}{r})*"
                        f"({get_column_letter(m + 4)}{r}>=DAY(DATE($B$2,{m + 1},0))))"
                        for m in range(1, 13)
                    )
                    qc.value = f"={done_parts}"
                elif fc == "W":
                    qc.value = f'=COUNTIF(E{r}:P{r},">=4")'
                else:
                    qc.value = f'=COUNTIF(E{r}:P{r},">=1")'
                qc.number_format = "0"

                # % Rate  (R = col 18)
                applicable_parts = "+".join(
                    f"(DATE($B$2,{m},1)<=TODAY())" for m in range(1, 13)
                )
                rc = ws.cell(row=r, column=18)
                rc.fill = fill_blue
                rc.alignment = ctr
                rc.border = thin_b
                rc.font = fnt_blueb
                rc.value = f'=IF(({applicable_parts})>0,Q{r}/({applicable_parts}),"")'
                rc.number_format = "0%"

                ws.row_dimensions[r].height = 20
                data_rows.append((r, fc))
                row_num += 1
                task_no += 1

    # ── Conditional Formatting — month cols E:P per row ──────────────
    # COLUMN()-4 gives month number (E=5→1, P=16→12)
    for r, fc in data_rows:
        rng = f"E{r}:P{r}"

        # Grey — future month
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f"DATE($B$2,COLUMN(E{r})-4,1)>TODAY()"],
                fill=fill_grey,
                font=fnt_grey,
            ),
        )

        # Decide required count formula
        if fc == "D":
            req = f"DAY(DATE($B$2,COLUMN(E{r})-3,0))"
        elif fc == "W":
            req = "4"
        else:
            req = "1"

        # Green — met target (cell has a number AND >= required)
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[
                    f"AND(DATE($B$2,COLUMN(E{r})-4,1)<=TODAY(),"
                    f"ISNUMBER(E{r}),E{r}>={req})"
                ],
                fill=fill_green,
                font=fnt_green,
            ),
        )

        # Red — not met (cell is blank OR < required, but month has passed)
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[
                    f"AND(DATE($B$2,COLUMN(E{r})-4,1)<=TODAY(),"
                    f"OR(NOT(ISNUMBER(E{r})),E{r}<{req}))"
                ],
                fill=fill_red,
                font=fnt_red,
            ),
        )

    # Timestamp footer
    ts = row_num + 1
    ws.merge_cells(start_row=ts, start_column=1, end_row=ts, end_column=9)
    ws.cell(
        row=ts, column=1
    ).value = f"Generated: {now.strftime('%d %b %Y  %H:%M')} (Malaysia time)"
    ws.cell(row=ts, column=1).font = Font(italic=True, size=8, color="AAAAAA")

    # ════════════════════════════════════════════════════════════════
    # SHEET 2 — TASKS REFERENCE
    # ════════════════════════════════════════════════════════════════
    ws_ref = wb.create_sheet("Tasks (Ref)")

    ws_ref.merge_cells("A1:D1")
    ws_ref["A1"].value = "Task Reference"
    ws_ref["A1"].fill = fill_hdr
    ws_ref["A1"].font = Font(color="FFFFFF", bold=True, size=11)
    ws_ref["A1"].alignment = ctr
    ws_ref.row_dimensions[1].height = 22

    for c, hdr in enumerate(["No.", "Location", "Frequency", "Phase"], 1):
        cell = ws_ref.cell(row=2, column=c)
        cell.value = hdr
        cell.fill = fill_hdr
        cell.font = fnt_hdr
        cell.alignment = ctr
        cell.border = thin_b
    ws_ref.row_dimensions[2].height = 18
    ws_ref.column_dimensions["A"].width = 6
    ws_ref.column_dimensions["B"].width = 44
    ws_ref.column_dimensions["C"].width = 12
    ws_ref.column_dimensions["D"].width = 8

    t_no = 1
    for task in tasks:
        phase = "P1" if "P1" in task["location"] else "P2"
        ws_ref.append([t_no, task["location"], task["frequency"], phase])
        rf = PatternFill("solid", fgColor="EBF4FF" if phase == "P1" else "E6FFF9")
        for c in range(1, 5):
            cell = ws_ref.cell(row=t_no + 2, column=c)
            cell.fill = rf
            cell.border = thin_b
            cell.alignment = ctr if c != 2 else lft
            cell.font = fnt_norm
        t_no += 1

    ws_ref.freeze_panes = "A3"
    wb.active = ws  # open on Checklist

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="Factory_Inspection_Checklist.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/tasks/add", methods=["GET", "POST"])
def add_task():
    if request.method == "POST":
        location = request.form.get("location", "").strip()
        frequency = request.form.get("frequency", "")
        task_type = request.form.get("type", "Internal")
        vendor = request.form.get("vendor", "").strip()

        if not location:
            flash("Location is required.", "error")
        elif frequency not in ALL_FREQUENCIES:
            flash("Invalid frequency.", "error")
        elif task_type not in ("Internal", "External"):
            flash("Invalid type.", "error")
        elif task_type == "External" and not vendor:
            flash("Vendor name is required for External tasks.", "error")
        else:
            vendor_val = vendor if task_type == "External" else None
            conn = get_db()
            conn.execute(
                "INSERT INTO tasks (location, frequency, type, vendor) VALUES (?, ?, ?, ?)",
                (location, frequency, task_type, vendor_val),
            )
            conn.commit()
            conn.close()
            flash(f'Task "{location}" added ({task_type}, {frequency}).', "success")
            return redirect(url_for("task_list"))
    return render_template("add_task.html", all_frequencies=ALL_FREQUENCIES)


@app.route("/tasks/delete/<int:task_id>", methods=["POST"])
def delete_task(task_id):
    conn = get_db()
    conn.execute("DELETE FROM task_logs WHERE task_id = ?", (task_id,))
    conn.execute("DELETE FROM tasks WHERE id = ?", (task_id,))
    conn.commit()
    conn.close()
    flash("Task deleted.", "success")
    return redirect(url_for("task_list"))


# ─────────────────────────────────────────────────────────────────
# Common Tasks routes
# ─────────────────────────────────────────────────────────────────


@app.route("/common-tasks")
def common_tasks():
    filter_status = request.args.get("status", "all")
    conn = get_db()
    if filter_status == "all":
        tasks = conn.execute(
            "SELECT * FROM common_tasks ORDER BY CASE status WHEN 'Open' THEN 1 WHEN 'In Progress' THEN 2 ELSE 3 END, "
            "CASE priority WHEN 'High' THEN 1 WHEN 'Medium' THEN 2 ELSE 3 END, created_date DESC"
        ).fetchall()
    else:
        tasks = conn.execute(
            "SELECT * FROM common_tasks WHERE status = ? ORDER BY "
            "CASE priority WHEN 'High' THEN 1 WHEN 'Medium' THEN 2 ELSE 3 END, created_date DESC",
            (filter_status,),
        ).fetchall()

    counts = conn.execute(
        "SELECT status, COUNT(*) as cnt FROM common_tasks GROUP BY status"
    ).fetchall()
    conn.close()

    count_map = {"Open": 0, "In Progress": 0, "Done": 0}
    for row in counts:
        count_map[row["status"]] = row["cnt"]

    return render_template(
        "common_tasks.html",
        tasks=tasks,
        filter_status=filter_status,
        count_map=count_map,
    )


@app.route("/common-tasks/add", methods=["GET", "POST"])
def add_common_task():
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "").strip()
        priority = request.form.get("priority", "Medium")
        if not title:
            flash("Title is required.", "error")
        elif priority not in ("High", "Medium", "Low"):
            flash("Invalid priority.", "error")
        else:
            now = get_malaysia_now()
            conn = get_db()
            conn.execute(
                "INSERT INTO common_tasks (title, description, priority, status, created_date) VALUES (?, ?, ?, ?, ?)",
                (
                    title,
                    description,
                    priority,
                    "Open",
                    now.strftime("%Y-%m-%d %H:%M:%S"),
                ),
            )
            conn.commit()
            conn.close()
            flash(f'Task "{title}" added!', "success")
            return redirect(url_for("common_tasks"))
    return render_template("add_common_task.html")


@app.route("/common-tasks/status/<int:task_id>", methods=["POST"])
def update_common_task_status(task_id):
    new_status = request.form.get("status", "")
    if new_status not in ("Open", "In Progress", "Done"):
        flash("Invalid status.", "error")
        return redirect(url_for("common_tasks"))

    now = get_malaysia_now()
    conn = get_db()
    if new_status == "Done":
        conn.execute(
            "UPDATE common_tasks SET status = ?, completed_date = ? WHERE id = ?",
            (new_status, now.strftime("%Y-%m-%d %H:%M:%S"), task_id),
        )
    else:
        conn.execute(
            "UPDATE common_tasks SET status = ?, completed_date = NULL WHERE id = ?",
            (new_status, task_id),
        )
    conn.commit()
    conn.close()
    flash("Task status updated.", "success")
    return redirect(url_for("common_tasks"))


@app.route("/common-tasks/delete/<int:task_id>", methods=["POST"])
def delete_common_task(task_id):
    conn = get_db()
    conn.execute("DELETE FROM common_tasks WHERE id = ?", (task_id,))
    conn.commit()
    conn.close()
    flash("Task deleted.", "success")
    return redirect(url_for("common_tasks"))


if __name__ == "__main__":
    import os

    init_db()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
